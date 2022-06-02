import openpyxl
import requests
import shutil
from openpyxl.drawing.image import Image
from PIL import Image as PILImage

wb = openpyxl.load_workbook(input('Nombre de archivo Excel con las incidencias: '))

ws = wb.active

dest_filename = input('Nombre de archivo a crear: ')

ws.delete_cols(1, 6)
ws.delete_cols(2, 1)
ws.delete_cols(3, 4)
ws.delete_cols(6, 8)
ws.delete_cols(8, 10)
ws.column_dimensions['H'].width = 51

for row in range(7, ws.max_row):
    ws.row_dimensions[row].height = 250

for row in range(7, ws.max_row):
    url = ws['H'+str(row)].value
    if url and not url.isspace(): 
        try:
            file_name = ws['D'+str(row)].value + ws['E'+str(row)].value + ".jpg"
            response = requests.get(url, stream = True)
            if response.status_code == 200:
                with open(file_name, 'wb') as f:
                    shutil.copyfileobj(response.raw, f)
                print('Imagen descargada: ',file_name)
            else:
                print('No se pudo descargar imagen')
                continue
            img = PILImage.open(ws['D'+str(row)].value + ws['E'+str(row)].value + ".jpg")
            (width, height) = (img.width // 3, img.height // 3)
            img_resized = img.resize((width, height))
            file_name_resized = ws['D'+str(row)].value + ws['E'+str(row)].value + "-resized.jpg"
            img_resized_file = img_resized.save(file_name_resized)
            img_excel = Image(file_name_resized)
            ws.add_image(img_excel, 'H'+ str(row))
            ws['H'+str(row)].value = ''
        except:
            print(response.status_code)

wb.save(filename = dest_filename)